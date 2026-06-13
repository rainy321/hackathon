# -*- coding: utf-8 -*-
"""
在已生成的 Growth_OS_规划.xlsx 上追加 4 个双人协作 sheet:
  11·团队分工
  12·协作解耦原则
  13·Agent任务清单(设计/PM 专属)
  14·48h双人计划
分工核心:技术=1人全栈;设计/PM=1人,但会用 Agent,
所以她承担“AI内容+体验+演示”这条与代码解耦的并行主线。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 配色(与主表一致) ----------
TITLE_BG    = "1F4E79"
SUB_BG      = "2E7D32"
TBL_BG      = "4472C4"
CARD_BLUE   = "DCE6F1"
CARD_GREEN  = "E2EFDA"
CARD_ORANGE = "FCE4D6"
CARD_PURPLE = "E6E0EC"   # 她/Agent 专属紫色卡片
CARD_GREY   = "F2F2F2"
WHITE       = "FFFFFF"
FONT_NAME   = "Microsoft YaHei"

def font(size=11, bold=False, color="000000"):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)
def fill(color):
    return PatternFill("solid", fgColor=color)
def align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(c, *, bold=False, size=11, color="000000", bg=None,
               wrap=True, h="left", v="top", border=True):
    c.font = font(size, bold, color)
    if bg: c.fill = fill(bg)
    c.alignment = align(wrap, h, v)
    if border: c.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def title_row(ws, text, ncols, row=1, height=34):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    style_cell(c, bold=True, size=16, color=WHITE, bg=TITLE_BG, h="left", v="center", border=False)
    set_row_height(ws, row, height)
def sub_title(ws, text, ncols, row, height=24):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    style_cell(c, bold=True, size=12, color=WHITE, bg=SUB_BG, h="left", v="center", border=False)
    set_row_height(ws, row, height)
def header_row(ws, headers, row, bg=TBL_BG):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        style_cell(c, bold=True, color=WHITE, bg=bg, h="center", v="center")
    set_row_height(ws, row, 22)
def data_rows(ws, rows, start_row, zebra=(WHITE, CARD_GREY)):
    r = start_row
    for i, row in enumerate(rows):
        for j in range(1, len(row) + 1):
            c = ws.cell(row=r, column=j, value=row[j-1])
            bg = zebra[i % 2]
            style_cell(c, bg=bg, v="center")
        maxlen = max(len(str(x)) for x in row)
        set_row_height(ws, r, max(20, min(140, 18 + (maxlen // 26) * 15)))
        r += 1
    return r

# =========================================================
path = r"C:\Users\zack\Desktop\hackathon\Growth_OS_规划.xlsx"
wb = openpyxl.load_workbook(path)

# 去重:若已存在同名 sheet 先删,保证可重复运行
for nm in ["11·团队分工", "12·协作解耦原则", "13·Agent任务清单", "14·48h双人计划"]:
    if nm in wb.sheetnames:
        del wb[nm]

# ---------------------------------------------------------
# Sheet 11: 团队分工
# ---------------------------------------------------------
ws = wb.create_sheet("11·团队分工")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [16, 40, 40, 22])
title_row(ws, "团队分工 · 2 人 / 48 小时", 4)

# 你
sub_title(ws, "👤 技术负责人(你)—— 全栈实现,独占代码", 4, 3)
header_row(ws, ["角色", "职责范围", "核心交付物", "方式/工具"], 4)
r = data_rows(ws, [[
    "技术\n(1人)",
    "架构设计 · 数据库 · 全部 API · 4 个 Prompt 接线 · 前端 · 联调修 Bug",
    "可跑通的 P0 闭环 + Dashboard/趋势图/记忆面板 UI",
    "Next.js + FastAPI + SQLite + LLM API",
]], 5, zebra=(CARD_BLUE, WHITE))

# 她
sub_title(ws, "🎨 设计 / 产品(她)—— 用 Agent 承担“AI 内容+体验+演示”主线", 4, r+1)
header_row(ws, ["角色", "职责范围", "核心交付物", "方式/工具"], r+2)
r = data_rows(ws, [[
    "设计/PM\n(1人)",
    "① Prompt 工程与调优(4 个核心 Prompt)\n② 种子数据 & Mock 响应(供前端先行)\n③ UI 设计与界面文案\n④ Demo 剧本 / Pitch / README\n⑤ 用户测试,记 Bug",
    "已验证的 Prompt + 示例输入输出 · Mock JSON · 界面稿 · 演示脚本 · Bug 清单",
    "Agent(Claude 等) · Figma / 手绘 · 文档工具",
]], r+3, zebra=(CARD_PURPLE, WHITE))

# 关键洞察
sub_title(ws, "💡 核心洞察:她不是“等着你给东西”的设计师", 4, r+1)
header_row(ws, ["要点", "说明", "", ""], r+2)
data_rows(ws, [
    ["不阻塞", "她用 Agent 自己造 Mock 数据/响应 → 你对着 stub 开发 → 最后接真 API。两人几乎全程并行。"],
    ["Prompt 归她", "Prompt 调得好不好决定产品质量,这是纯“内容”工作,不需要写代码,最适合她用 Agent 迭代。"],
    ["Demo 归她", "评委看的是故事和系统感,她出 Demo 剧本+Pitch,价值与你的代码同等重要。"],
    ["你不孤单", "代码全你写,但她负责“喂好 AI”和“讲好故事”,你专注把管道跑通即可。"],
], r+3)

# ---------------------------------------------------------
# Sheet 12: 协作解耦原则
# ---------------------------------------------------------
ws = wb.create_sheet("12·协作解耦原则")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [22, 50, 30])
title_row(ws, "协作解耦原则 · 让两人不互相阻塞", 3)
header_row(ws, ["原则", "怎么做", "谁负责"], 2)
data_rows(ws, [
    ["1. 先定数据契约", "开工 2 小时内,她用 Agent 输出“API 路由 + 字段 + Mock JSON”,你照着建库建表", "她出契约 / 你实现"],
    ["2. 前端对 Mock 先行", "她持续产出各接口的假返回,前端不等后端就能把界面做出来,后端就绪再切真", "她造 Mock / 你切真"],
    ["3. Prompt 写测分离", "她写并跑测试用例调 Prompt,你只做“接线”(把她的 Prompt 贴进 AI Adapter)", "她写测 / 你接线"],
    ["4. 明确交接物", "每个里程碑约定具体交付物(如“阶段1末交 Prompt+示例IO”),到期对齐", "双方"],
    ["5. 单一信息源", "用一份在线文档(字段表 / Mock / Prompt / Bug)做唯一来源,避免口头传话出错", "她维护 / 你同步"],
    ["6. 每日 3 次同步", "早/中/晚各 10 分钟对齐进度与阻塞,其余时间各自并行", "双方"],
], 3)

# ---------------------------------------------------------
# Sheet 13: Agent 任务清单(设计/PM 专属)
# ---------------------------------------------------------
ws = wb.create_sheet("13·Agent任务清单")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [20, 24, 24, 34, 12])
title_row(ws, "Agent 任务清单 · 设计/PM 用 Agent 可独立完成的工作", 5)
header_row(ws, ["任务", "用 Agent 怎么做", "输入", "交付物", "对应阶段"], 2)
data_rows(ws, [
    ["API 契约 + Mock", "让 Agent 按 MVP 生成路由/字段/示例 JSON", "数据模型 + MVP 范围", "API 文档 + Mock JSON", "0"],
    ["Prompt 工程 ×4", "写 Goal Decomposer / Daily Planner / Behavior Analyst / System Adjuster,跑用例调优",
     "目标样本 + 行为样本", "4 个终版 Prompt + 示例输入输出", "0–1"],
    ["Prompt 测试用例", "造多组输入(易/难/边界)看输出是否落地、是否稳定", "Prompt", "测试报告(哪条该改)", "1–3"],
    ["种子数据", "生成示例用户/目标/任务/行为日志,供 Demo 演示", "场景设定", "一份真实感的演示数据", "1–2"],
    ["UI 设计", "用 Agent 描述 Dashboard/打卡/记忆面板信息架构,再手绘或 Figma", "核心闭环 + 模块清单", "界面稿/线框", "0–2"],
    ["界面文案", "Empty State、提示、按钮、记忆面板“系统总结”话术", "界面稿", "完整文案表", "2–3"],
    ["Demo 剧本", "设计 3 个典型用户故事(如“考研党/健身/学英语”)串成演示动线", "种子数据", "演示脚本(带旁白)", "3–4"],
    ["Pitch / README", "用 Agent 起草项目介绍、卖点、架构图说明", "本项目规划", "Pitch 文档 + README", "3–5"],
    ["用户测试", "像真实用户跑全流程,记录体验问题与 Bug", "可跑通的产品", "Bug 优先级清单", "4"],
], 3)

# ---------------------------------------------------------
# Sheet 14: 48h 双人计划
# ---------------------------------------------------------
ws = wb.create_sheet("14·48h双人计划")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [9, 9, 40, 40, 26])
title_row(ws, "48 小时双人计划 · 技术(你) ‖ 设计/PM·Agent驱动(她)", 5)
header_row(ws, ["阶段", "时间", "技术线 · 你", "设计/PM 线 · 她(Agent)", "交接 / 同步点"], 2)
data_rows(ws, [
    ["阶段0\n对齐起骨架", "0–4h",
     "确认数据模型+API契约 → 搭 FastAPI+SQLite+Next.js 骨架 → 建空表",
     "用 Agent 出 API 契约+Mock JSON;定视觉风格与信息架构",
     "2h:她交契约 → 你建库建表"],
    ["阶段1\n闭环+Prompt", "4–12h",
     "Goal/Task/Log/Memory 的 CRUD → 搭 AI Adapter(等她的 Prompt 接线)",
     "写并调优 4 个 Prompt,跑测试用例 → 产出“终版 Prompt+示例IO”;出 Dashboard/打卡设计",
     "末:她交 Prompt+示例IO → 你接线"],
    ["阶段2\n打通主流程", "12–24h",
     "打通“拆解→每日任务→打卡”P0 主流程;前端对接",
     "持续推进界面落地(对着 Mock);设计时间视图/趋势图;写界面文案",
     "中:后端真 API 就绪 → 前端切真数据"],
    ["阶段3\n反馈系统", "24–34h",
     "行为分析 + 动态调整 + Memory 闭环",
     "用真实数据测 Prompt 输出质量并反馈;优化记忆面板与“系统总结”话术",
     "持续:她反馈 Prompt 问题 → 你调"],
    ["阶段4\nUI+Demo", "34–44h",
     "Dashboard/趋势图/记忆面板落地;按她的 Bug 清单修 Bug",
     "录 Demo 剧本+Pitch+README;用户测试全流程,出优先级 Bug 清单",
     "末:她交 Bug 清单 → 你修"],
    ["阶段5\n收尾彩排", "44–48h",
     "联调 / 性能 / 收尾 Bug / 部署可演示环境",
     "最终演示彩排;Pitch/PPT 定稿;准备评委问答",
     "终:联合彩排一遍"],
], 3)

# 关键提醒
sub_title(ws, "⚠️ 关键提醒", 5, 11)
data_rows(ws, [
    ["并行", "两人绝不串行等待:她永远有 Mock/内容/设计可做,你永远有接口/前端可写。"],
    ["Prompt 是产品", "Prompt 调得好 = 演示惊艳,这是她用 Agent 能创造的最高价值,优先投入。"],
    ["Demo 先行", "44h 前必须冻结功能,留足彩排时间;别在最后 2 小时还在加新功能。"],
    ["单一文档", "字段/Mock/Prompt/Bug 全放一份共享文档,实时更新。"],
], 12)

import os
def save_anywhere(wb, path):
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        alt = f"{base}_新{ext}"
        wb.save(alt)
        return alt

saved = save_anywhere(wb, path)
print("OK saved:", saved)
print("Sheets:", wb.sheetnames)
