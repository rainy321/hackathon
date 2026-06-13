# -*- coding: utf-8 -*-
"""
生成 Growth OS(长期成长操作系统)黑客松规划 Excel。
基于需求文档,拆解为多个结构化 sheet,便于展示与开发规划。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 配色 ----------
TITLE_BG    = "1F4E79"   # 深蓝 - 大标题
SUB_BG      = "2E7D32"   # 深绿 - 区块小标题
TBL_BG      = "4472C4"   # 表头蓝
CARD_BLUE   = "DCE6F1"   # 浅蓝卡片
CARD_GREEN  = "E2EFDA"   # 浅绿卡片
CARD_ORANGE = "FCE4D6"   # 浅橙卡片(P2)
CARD_GREY   = "F2F2F2"   # 浅灰
WHITE       = "FFFFFF"
FONT_NAME   = "Microsoft YaHei"

# ---------- 样式工厂 ----------
def font(size=11, bold=False, color="000000"):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(c, *, bold=False, size=11, color="000000", bg=None,
               wrap=True, h="left", v="top", border=True):
    c.font = font(size, bold, color)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(wrap, h, v)
    if border:
        c.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def title_row(ws, text, ncols, row=1, height=34):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    style_cell(c, bold=True, size=16, color=WHITE, bg=TITLE_BG, h="left", v="center", border=False)
    set_row_height(ws, row, height)

def sub_title(ws, text, ncols, row, height=24):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    style_cell(c, bold=True, size=12, color=WHITE, bg=SUB_BG, h="left", v="center", border=False)
    set_row_height(ws, row, height)

def header_row(ws, headers, row, bg=TBL_BG):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        style_cell(c, bold=True, color=WHITE, bg=bg, h="center", v="center")
    set_row_height(ws, row, 22)

def data_rows(ws, rows, start_row, zebra=(WHITE, CARD_GREY), min_col=None):
    r = start_row
    for i, row in enumerate(rows):
        cols = min_col or len(row)
        for j in range(1, cols + 1):
            val = row[j-1] if j-1 < len(row) else ""
            c = ws.cell(row=r, column=j, value=val)
            bg = zebra[i % 2]
            style_cell(c, bg=bg, v="center")
        # 行高自适应(粗略)
        maxlen = max((len(str(row[j-1])) if j-1 < len(row) else 0) for j in range(1, cols+1))
        set_row_height(ws, r, max(20, min(120, 18 + (maxlen // 28) * 15)))
        r += 1
    return r

def kv_card(ws, pairs, start_row, k_w_bg=CARD_BLUE, height=None):
    """属性-值 两列卡片,合并到 sheet 总列数=2"""
    r = start_row
    for k, v in pairs:
        kc = ws.cell(row=r, column=1, value=k)
        vc = ws.cell(row=r, column=2, value=v)
        style_cell(kc, bold=True, bg=k_w_bg, h="center", v="center")
        style_cell(vc, bg=WHITE, v="center")
        ln = len(str(v))
        set_row_height(ws, r, max(22, min(90, 20 + (ln // 40) * 16)))
        r += 1
    return r

# =========================================================
wb = Workbook()

# ---------------------------------------------------------
# Sheet 1: 项目概览
# ---------------------------------------------------------
ws = wb.active
ws.title = "1·项目概览"
ws.sheet_view.showGridLines = False
set_col_widths(ws, [22, 95])
title_row(ws, "Growth OS · 长期成长操作系统(黑客松规划)", 2, height=40)

sub_title(ws, "一、项目定位", 2, 3)
r = kv_card(ws, [
    ("项目名称", "Growth OS — 长期成长操作系统(LifeOS / Growth OS)"),
    ("一句话定义", "一个“把人生拆成可计算系统”的长期成长操作系统"),
    ("本质定位", "黑客松里偏“产品级系统设计”的项目,而非技术炫技"),
    ("核心理念", "时间 = 目标复利系统(目标 → 能力 → 训练 → 每日动作)"),
    ("目标用户", "想实现长期成长复利的人:学习 / 健康 / 习惯 / 目标管理"),
], 4)

sub_title(ws, "二、核心数据(产品骨架)", 2, r+1)
r = kv_card(ws, [
    ("状态维度", "4 个:🧠学习维度  💪健康维度  ⚙️行为维度  🎯目标维度"),
    ("时间结构", "5 层:Year Goal → Quarter → Month → Week → Daily Tasks"),
    ("三大能力", "A.目标拆解  B.行为监督  C.反馈系统(系统反过来改计划)"),
    ("核心闭环", "设目标→AI拆解→生成每日任务→执行→记录行为→分析偏差→动态调整→循环"),
], r+2, k_w_bg=CARD_GREEN)

sub_title(ws, "三、技术与 AI 策略", 2, r+1)
r = kv_card(ws, [
    ("技术栈", "Next.js(前端) + FastAPI(API层) + SQLite(数据) + LLM API(AI Adapter)"),
    ("AI 设计原则", "✅ 单 Agent + 4 个 Prompt 函数;❌ 不用 LangChain/AutoGen/CrewAI/多Agent框架"),
    ("不做清单(P2)", "多Agent、向量数据库、知识图谱、MCP、自动化执行系统"),
    ("MVP 周期", "48 小时内做出 P0 + 部分 P1"),
], r+2, k_w_bg=CARD_BLUE)

# ---------------------------------------------------------
# Sheet 2: 产品设计
# ---------------------------------------------------------
ws = wb.create_sheet("2·产品设计")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [16, 30, 60])
title_row(ws, "产品设计 · 多维建模 / 时间结构 / 核心能力", 3)

sub_title(ws, "A. 多维人生建模(4 个状态维度)", 3, 3)
header_row(ws, ["维度", "包含项", "说明"], 4)
r = data_rows(ws, [
    ["🧠 学习维度", "认知 / 技能 / 表达", "知识储备与能力成长"],
    ["💪 健康维度", "饮食 / 运动 / 睡眠", "身体状态与能量水平"],
    ["⚙️ 行为维度", "习惯 / 执行力 / 纪律", "能否持续把动作做下去"],
    ["🎯 目标维度", "长期目标 / 阶段目标", "给前三个维度提供方向"],
], 5)

sub_title(ws, "B. 时间结构(目标复利系统,非任务软件)", 3, r+1)
header_row(ws, ["层级", "名称", "作用"], r+2)
r = data_rows(ws, [
    ["L1", "Year Goal 年目标", "长期方向(1年)"],
    ["L2", "Quarter Goal 季度目标", "拆成阶段(3个月)"],
    ["L3", "Month Plan 月计划", "本月的训练重点"],
    ["L4", "Week Plan 周计划", "本周要完成的具体块"],
    ["L5", "Daily Tasks 每日任务", "今天可执行的动作(≤5个)"],
], r+3)

sub_title(ws, "C. 三大核心能力", 3, r+1)
header_row(ws, ["能力", "描述", "输入 → 输出"], r+2)
r = data_rows(ws, [
    ["A. 目标拆解\nGoal Decomposition", "把模糊目标变成可落地的每日动作",
     "输入:“我要变强/考研/健身/学英语”\n输出:长期目标→能力拆解→训练路径→每日动作"],
    ["B. 行为监督\nBehavior Tracking", "记录不是打卡,而是关注质量、连续性、波动原因",
     "记录:有没有做 / 做的质量(1-5) / 连续天数 / 中断原因"],
    ["C. 反馈系统\nFeedback(最关键)", "系统会反过来改计划,形成自适应闭环",
     "示例:连续没学英语→降难度;运动过量→自动调休;效率低→换学习方式"],
], r+3)

# ---------------------------------------------------------
# Sheet 3: 系统架构
# ---------------------------------------------------------
ws = wb.create_sheet("3·系统架构")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [16, 24, 56])
title_row(ws, "系统架构 · 分层结构", 3)
header_row(ws, ["层级", "组件", "职责"], 2)
data_rows(ws, [
    ["前端层", "Next.js", "Dashboard / 目标输入 / 每日任务 / 打卡 / 趋势图 / 记忆面板"],
    ["API 层", "FastAPI", "REST 接口,串联前端与核心引擎,调用 LLM"],
    ["核心引擎", "Goal Engine 目标引擎", "目标拆解与能力建模"],
    ["核心引擎", "Planner 计划器", "根据目标+状态+历史生成每日任务"],
    ["核心引擎", "Tracker 追踪器", "行为记录、完成率、连续性统计"],
    ["核心引擎", "Memory 长期记忆", "沉淀用户洞察(如“晚上效率更高”)"],
    ["核心引擎", "AI Adapter", "封装 LLM 调用,对接 4 个 Prompt 函数"],
    ["数据层", "SQLite", "User / Goal / Task / Behavior Log / Memory"],
], 3)

# ---------------------------------------------------------
# Sheet 4: 数据模型
# ---------------------------------------------------------
ws = wb.create_sheet("4·数据模型")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [26, 18, 50])
title_row(ws, "核心数据模型(5 张表)", 3)

def table_block(ws, name, fields, start_row, bg):
    sub_title(ws, name, 3, start_row)
    header_row(ws, ["字段", "类型", "说明"], start_row+1)
    return data_rows(ws, fields, start_row+2)

r = 3
r = table_block(ws, "① User 用户表", [
    ["id", "int / PK", "用户唯一ID"],
    ["name", "str", "用户名称"],
    ["baseline", "json", "当前状态基线(各维度起点)"],
], r, CARD_BLUE) + 1

r = table_block(ws, "② Goal 目标表", [
    ["id", "int / PK", "目标唯一ID"],
    ["user_id", "int / FK", "所属用户"],
    ["title", "str", "目标标题(如:减肥10kg)"],
    ["category", "enum", "health / study / behavior / goal"],
    ["time_horizon", "str", "时间跨度(如 3 个月)"],
    ["status", "enum", "进行中 / 已完成 / 已暂停"],
], r, CARD_GREEN) + 1

r = table_block(ws, "③ Task 任务表(执行单元)", [
    ["id", "int / PK", "任务唯一ID"],
    ["goal_id", "int / FK", "所属目标"],
    ["date", "date", "执行日期"],
    ["content", "str", "今天要做什么"],
    ["difficulty", "int 1-5", "难度等级"],
    ["status", "enum", "待办 / 已完成 / 已跳过"],
], r, CARD_BLUE) + 1

r = table_block(ws, "④ Behavior Log 行为反馈表", [
    ["id", "int / PK", "记录唯一ID"],
    ["task_id", "int / FK", "关联任务"],
    ["done", "0 / 1", "有没有做"],
    ["quality", "int 1-5", "做的质量"],
    ["notes", "str", "备注 / 中断原因"],
], r, CARD_GREEN) + 1

r = table_block(ws, "⑤ Memory 长期记忆表", [
    ["id", "int / PK", "记忆唯一ID"],
    ["user_id", "int / FK", "所属用户"],
    ["insight_type", "str", "洞察类型"],
    ["content", "str", "例:“用户晚上效率更高”“易在第3天断掉习惯”“偏好短任务”"],
], r, CARD_BLUE) + 1

# ---------------------------------------------------------
# Sheet 5: AI Prompt 设计
# ---------------------------------------------------------
ws = wb.create_sheet("5·AI Prompt 设计")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [18, 14, 24, 24, 44])
title_row(ws, "AI Prompt 设计 · 单 Agent + 4 个核心函数", 5)
header_row(ws, ["Prompt 函数", "角色", "输入", "输出要求", "提示词要点"], 2)
data_rows(ws, [
    ["① Goal Decomposer\n目标拆解", "人生规划教练",
     "用户原始目标",
     "长期能力 / 3阶段路径 / 可执行任务",
     "你是一个人生规划教练。把用户目标拆解为:长期目标→能力拆解→训练路径→每日动作。要求:必须可落地到每日任务。"],
    ["② Daily Planner\n每日生成", "每日计划员",
     "目标 + 当前状态 + 历史行为",
     "今天的任务清单",
     "根据用户目标+当前状态+历史行为生成今天任务。要求:不超过5个任务、有优先级、可执行。"],
    ["③ Behavior Analyst\n行为分析", "行为分析师",
     "过去7天行为数据",
     "规律 / 问题 / 调整建议",
     "分析用户过去7天行为,输出:规律、问题、建议调整。"],
    ["④ System Adjuster\n动态调整", "计划调整器",
     "行为数据 + 当前计划",
     "降/升难度,改任务类型",
     "根据行为数据调整目标计划:降难度 / 提升难度 / 改任务类型。"],
], 3)

# ---------------------------------------------------------
# Sheet 6: 核心闭环
# ---------------------------------------------------------
ws = wb.create_sheet("6·核心闭环")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [8, 20, 40, 22, 22])
title_row(ws, "核心闭环 · 产品灵魂(复利系统)", 5)
header_row(ws, ["步骤", "环节", "说明", "输入", "输出"], 2)
data_rows(ws, [
    ["1", "用户设定目标", "用户输入想做的事", "原始目标", "Goal 记录"],
    ["2", "AI 拆解目标", "调用 Goal Decomposer", "Goal", "能力 + 训练路径"],
    ["3", "生成每日任务", "调用 Daily Planner", "目标+状态", "Daily Tasks"],
    ["4", "用户执行", "完成每日任务", "Daily Tasks", "实际行为"],
    ["5", "记录行为数据", "Tracker 记录 done/quality", "执行情况", "Behavior Log"],
    ["6", "AI 分析偏差", "调用 Behavior Analyst", "7天行为", "规律/问题/建议"],
    ["7", "自动调整计划", "调用 System Adjuster", "分析+当前计划", "新计划"],
    ["8", "循环(复利)", "回到第 3 步,持续迭代", "新计划", "成长复利"],
], 3)

# ---------------------------------------------------------
# Sheet 7: MVP 功能拆解
# ---------------------------------------------------------
ws = wb.create_sheet("7·MVP功能拆解")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [10, 22, 50, 12])
title_row(ws, "MVP 功能拆解(P0 / P1 / P2)", 4)
header_row(ws, ["优先级", "功能", "描述", "是否做"], 2)
p0 = [
    ["P0 必须", "目标输入", "用户输入想实现的目标", "✅ 做"],
    ["P0 必须", "自动拆解", "AI 把目标拆成能力+路径", "✅ 做"],
    ["P0 必须", "每日任务生成", "生成今天 ≤5 个任务", "✅ 做"],
    ["P0 必须", "打卡", "记录任务完成情况", "✅ 做"],
    ["P0 必须", "简单反馈", "完成率 / 连续天数", "✅ 做"],
]
p1 = [
    ["P1 加分", "简单趋势图", "执行稳定性 / 能力曲线", "⭐ 加分"],
    ["P1 加分", "AI 总结行为", "Behavior Analyst 输出", "⭐ 加分"],
]
p2 = [
    ["P2 不做", "多Agent", "复杂度爆炸,放弃", "❌ 不做"],
    ["P2 不做", "向量数据库", "MVP 不需要", "❌ 不做"],
    ["P2 不做", "知识图谱", "过度工程化", "❌ 不做"],
    ["P2 不做", "MCP", "超出范围", "❌ 不做"],
    ["P2 不做", "自动化执行系统", "超出范围", "❌ 不做"],
]
r = 3
r = data_rows(ws, p0, r, zebra=(CARD_GREEN, WHITE))
r = data_rows(ws, p1, r, zebra=(CARD_BLUE, WHITE))
r = data_rows(ws, p2, r, zebra=(CARD_ORANGE, WHITE))

# ---------------------------------------------------------
# Sheet 8: API 设计
# ---------------------------------------------------------
ws = wb.create_sheet("8·API 设计")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [10, 30, 30, 32, 28])
title_row(ws, "API 设计 · REST 接口", 5)
header_row(ws, ["方法", "路径", "描述", "请求体(示例)", "响应(示例)"], 2)
data_rows(ws, [
    ["POST", "/api/goals", "创建并 AI 拆解目标",
     '{"title":"减肥10kg","category":"health","time_horizon":"3M"}',
     '{"goal_id":1,"decomposition":[...]}'],
    ["GET",  "/api/goals", "获取用户目标列表", "—", "[{goal_id, title, progress}]"],
    ["POST", "/api/tasks/today", "生成今日任务(Daily Planner)",
     '{"goal_id":1,"state":{...}}',
     '{"tasks":[{content,difficulty,priority}]}'],
    ["GET",  "/api/tasks/today", "获取今日任务", "—", "[{task_id, content, status}]"],
    ["POST", "/api/tasks/{id}/checkin", "打卡 / 记录行为",
     '{"done":1,"quality":4,"notes":"..."}',
     '{"task_id":3,"streak":7}'],
    ["GET",  "/api/analysis", "行为分析(过去7天)",
     '?days=7', '{"patterns":[],"issues":[],"advice":[]}'],
    ["POST", "/api/adjust", "动态调整计划",
     '{"goal_id":1,"analysis":{...}}',
     '{"adjustments":[{type:"lower_difficulty"}]}'],
    ["GET",  "/api/memory", "获取长期记忆/洞察", "—", "[{insight_type, content}]"],
    ["GET",  "/api/dashboard", "Dashboard 聚合数据", "—", '{"completion_rate","streak","progress"}'],
], 3)

# ---------------------------------------------------------
# Sheet 9: 可视化设计
# ---------------------------------------------------------
ws = wb.create_sheet("9·可视化设计")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [20, 50, 28])
title_row(ws, "可视化设计 · 强调“系统感”(黑客松加分)", 3)
header_row(ws, ["模块", "内容", "展示形式"], 2)
data_rows(ws, [
    ["Dashboard 仪表盘", "今日完成率 / 连续天数 / 能力成长", "数字卡片 + 进度环"],
    ["目标进度", "学习 67% / 健康 40% 等各维度进度", "进度条 / 雷达图"],
    ["趋势图", "执行稳定性 / 能量曲线", "折线图"],
    ["时间视图", "Today → Week → Month → Quarter 切换", "时间轴 / Tab 切换"],
    ["记忆面板", "系统总结的洞察(晚上效率高/易断习惯…)", "卡片列表"],
], 3)

# ---------------------------------------------------------
# Sheet 10: 开发时间线(48h)
# ---------------------------------------------------------
ws = wb.create_sheet("10·开发时间线")
ws.sheet_view.showGridLines = False
set_col_widths(ws, [10, 14, 40, 28])
title_row(ws, "开发时间线 · 48 小时分配(参考)", 4)
header_row(ws, ["阶段", "时间", "任务", "产出"], 2)
data_rows(ws, [
    ["阶段0", "0–4h", "需求确认 + 数据库设计 + 搭骨架", "SQLite + FastAPI/Next.js 脚手架"],
    ["阶段1", "4–12h", "核心数据 CRUD + 4 个 Prompt 函数", "Goal/Task/Log/Memory 接口 + AI Adapter"],
    ["阶段2", "12–24h", "打通核心闭环(拆解→每日任务→打卡)", "可跑通的 P0 主流程"],
    ["阶段3", "24–34h", "行为分析 + 动态调整 + 记忆", "反馈系统闭环"],
    ["阶段4", "34–44h", "前端 Dashboard / 趋势图 / 记忆面板", "系统感 UI"],
    ["阶段5", "44–48h", "联调 / 录 Demo / 写文档", "可演示成品"],
], 3)

# ---------- 保存 ----------
out = r"C:\Users\zack\Desktop\hackathon\Growth_OS_规划.xlsx"
wb.save(out)
print("已生成:", out)
print("Sheets:", wb.sheetnames)
